"""
verify_gene_table.py
=====================
Verifies "Table 1" of a gene-data workbook (like Genes_10th.xlsx) against
public databases and internal consistency rules, and writes an annotated
report workbook.

WHAT IT CHECKS
--------------
Offline (no internet needed):
  - ID fields with stray whitespace / newlines (common copy-paste artifact)
  - CDS_Sequence length vs stated CDS_Length_bp
  - Full_Sequence length vs stated Full_Sequence_Length_pb
  - CDS_Sequence is a substring of Full_Sequence
  - CDS translates (in-frame) to the stated Protein_Sequence
  - Protein_Sequence length vs stated AA_Length
  - Non-standard amino acid characters in Protein_Sequence
  - Molecular weight recomputed from Protein_Sequence vs stated MW_kDa
  - Domain_End vs AA_Length sanity

Online (needs internet + NCBI/UniProt/InterPro access):
  - NCBI_Accession resolves on NCBI Nucleotide, and its sequence matches
    Full_Sequence (falls back to comparing against CDS_Sequence if the
    accession is CDS-only)
  - Protein_accession resolves on NCBI Protein, and its sequence matches
    Protein_Sequence
  - Protein_accession maps to a UniProt entry, and that entry lists the
    stated InterPro_Domain (IPR id) -- best effort; not every RefSeq
    protein has a UniProt cross-reference
  - NCBI_Accession's record title mentions the stated Gene_Name or
    Abbreviation (loose substring match, since NCBI's phrasing won't be
    identical text), and the record's organism matches the stated Species

SETUP
-----
    pip install biopython requests openpyxl pandas

Edit the CONFIG block below (your email is required by NCBI Entrez;
an NCBI API key is optional but raises your rate limit from 3 to 10
requests/second -- get one free at
https://www.ncbi.nlm.nih.gov/account/settings/).

RUN
---
    python verify_gene_table.py

OUTPUT
------
    gene_table_verification_report.xlsx
        - "Table1_Annotated": your original rows + one status column per
          check (PASS / FAIL / WARN / SKIP) + a Detail column
        - "Issues_Only": just the rows/fields that failed or warned

NOTES / LIMITATIONS
--------------------
  - Locus IDs (Gene_Locus_ID) are NOT automatically re-verified: your
    dataset mixes AT-locus tags, LOC ids, cultivar-specific tags, and
    free-text identifiers across 4 species, and there's no single
    database endpoint that reliably resolves all of those formats. The
    NCBI accession checks are the reliable cross-check instead.
  - InterPro matching depends on a RefSeq->UniProt mapping existing;
    if it doesn't, that row is marked SKIP, not FAIL.
  - This is a research aid, not a source of truth: always read the
    Detail column before trusting a FAIL.
"""

import re
import time
import sys

import pandas as pd
import requests
from Bio.Seq import Seq
from Bio.SeqUtils.ProtParam import ProteinAnalysis

# ============================== CONFIG ===============================
INPUT_XLSX = "Genes_10th.xlsx"          # path to your workbook
SHEET_NAME = "Table 1"                  # only this sheet is used
OUTPUT_XLSX = "gene_table_verification_report.xlsx"

NCBI_EMAIL = "you@example.com"          # REQUIRED by NCBI Entrez -- put your real email
NCBI_API_KEY = None                     # optional, e.g. "abcdef123456..."

RUN_ONLINE_CHECKS = True                # set False to only run offline checks
REQUEST_DELAY = 0.34 if not NCBI_API_KEY else 0.11   # seconds between NCBI calls (3/s or 10/s)
MW_TOLERANCE_PCT = 1.0                  # allowed %% difference before flagging MW mismatch
# =======================================================================

STD_AA = set("ACDEFGHIKLMNPQRSTVWY")
NCBI_EUTILS = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
UNIPROT_IDMAP = "https://rest.uniprot.org/idmapping"
UNIPROT_ENTRY = "https://rest.uniprot.org/uniprotkb"


def clean(x):
    """Strip whitespace/newlines that show up as copy-paste artifacts."""
    if pd.isna(x):
        return None
    return re.sub(r"\s+", "", str(x))


def had_whitespace(x):
    if pd.isna(x):
        return False
    s = str(x)
    return s != s.strip() or "\n" in s or "\r" in s


# ------------------------- offline checks -------------------------

def run_offline_checks(row):
    results = {}

    for col in ["Gene_Locus_ID", "NCBI_Accession", "Protein_accession", "InterPro_Domain"]:
        if had_whitespace(row[col]):
            results[f"{col}_format"] = ("WARN", f"stray whitespace/newline in value: {row[col]!r}")

    cds = clean(row["CDS_Sequence"])
    full = clean(row["Full_Sequence"]) if not pd.isna(row["Full_Sequence"]) else None
    prot_stated = clean(row["Protein_Sequence"])

    # CDS length
    if cds and not pd.isna(row["CDS_Length_bp"]):
        stated = int(row["CDS_Length_bp"])
        results["CDS_Length_bp"] = ("PASS", "") if len(cds) == stated else (
            "FAIL", f"stated {stated}, actual {len(cds)}")

    # Full sequence length
    if full and not pd.isna(row["Full_Sequence_Length_pb"]):
        stated = int(row["Full_Sequence_Length_pb"])
        results["Full_Sequence_Length_pb"] = ("PASS", "") if len(full) == stated else (
            "FAIL", f"stated {stated}, actual {len(full)}")

    # CDS substring of full sequence
    if cds and full:
        results["CDS_in_Full_Sequence"] = ("PASS", "") if cds in full else (
            "WARN", "CDS not found as a contiguous substring of Full_Sequence")

    # Translation check
    translated = None
    if cds:
        if len(cds) % 3 != 0:
            results["CDS_Translation"] = ("FAIL", f"CDS length {len(cds)} not divisible by 3")
        else:
            try:
                translated = str(Seq(cds).translate(to_stop=True))
            except Exception as e:
                results["CDS_Translation"] = ("FAIL", f"translation error: {e}")
    if translated is not None and prot_stated:
        results["CDS_Translation"] = ("PASS", "") if translated == prot_stated else (
            "FAIL", f"translated protein (len {len(translated)}) != stated Protein_Sequence (len {len(prot_stated)})")

    # AA length
    if prot_stated and not pd.isna(row["AA_Length"]):
        stated = int(row["AA_Length"])
        results["AA_Length"] = ("PASS", "") if len(prot_stated) == stated else (
            "FAIL", f"stated {stated}, actual {len(prot_stated)}")

    # non-standard chars
    if prot_stated:
        bad = sorted(set(prot_stated) - STD_AA)
        if bad:
            results["Protein_Sequence_chars"] = ("WARN", f"non-standard characters: {bad}")

    # MW recompute
    if prot_stated and not pd.isna(row["MW_kDa"]):
        seq_for_mw = prot_stated.replace("*", "")
        bad = set(seq_for_mw) - STD_AA
        if bad:
            results["MW_kDa"] = ("SKIP", f"non-standard residues present {sorted(bad)}, skipping MW calc")
        else:
            mw = ProteinAnalysis(seq_for_mw).molecular_weight() / 1000.0
            stated = float(row["MW_kDa"])
            pct = abs(mw - stated) / stated * 100 if stated else 999
            if pct > MW_TOLERANCE_PCT:
                results["MW_kDa"] = ("FAIL", f"stated {stated:.2f} kDa, recalculated {mw:.2f} kDa ({pct:.1f}% diff)")
            else:
                results["MW_kDa"] = ("PASS", "")

    # domain range sanity
    if not pd.isna(row["Domain_End"]) and not pd.isna(row["AA_Length"]):
        de, aal = int(row["Domain_End"]), int(row["AA_Length"])
        results["Domain_Range"] = ("PASS", "") if de <= aal else (
            "FAIL", f"Domain_End {de} > AA_Length {aal}")

    return results


# ------------------------- online checks -------------------------

def entrez_params(extra):
    p = {"email": NCBI_EMAIL, "tool": "gene_table_verifier"}
    if NCBI_API_KEY:
        p["api_key"] = NCBI_API_KEY
    p.update(extra)
    return p


def fetch_ncbi_summary(accession, db):
    """db = 'nuccore' or 'protein'. Returns dict with 'title' and 'organism', or None."""
    if not accession:
        return None
    try:
        # esearch first to resolve accession -> internal UID (esummary needs a UID, not always the accession)
        r = requests.get(f"{NCBI_EUTILS}/esearch.fcgi", params=entrez_params({
            "db": db, "term": accession, "retmode": "json"
        }), timeout=20)
        time.sleep(REQUEST_DELAY)
        ids = r.json().get("esearchresult", {}).get("idlist", [])
        if not ids:
            return None
        uid = ids[0]
        r = requests.get(f"{NCBI_EUTILS}/esummary.fcgi", params=entrez_params({
            "db": db, "id": uid, "retmode": "json"
        }), timeout=20)
        time.sleep(REQUEST_DELAY)
        result = r.json().get("result", {})
        doc = result.get(uid, {})
        return {
            "title": doc.get("title", ""),
            "organism": doc.get("organism", ""),
        }
    except Exception:
        return None


def name_matches_title(gene_name, abbreviation, title):
    """Loose check: does the gene name or abbreviation show up in the NCBI title?"""
    title_norm = re.sub(r"[^a-z0-9]", "", title.lower())
    for candidate in [gene_name, abbreviation]:
        if not candidate:
            continue
        cand_norm = re.sub(r"[^a-z0-9]", "", str(candidate).lower())
        if cand_norm and cand_norm in title_norm:
            return True
    return False


def species_matches_organism(species, organism):
    if not species or not organism:
        return False
    return species.strip().lower() == organism.strip().lower()


def fetch_ncbi_sequence(accession, db):
    """db = 'nuccore' or 'protein'. Returns raw sequence string (no header) or None."""
    if not accession:
        return None
    try:
        r = requests.get(f"{NCBI_EUTILS}/efetch.fcgi", params=entrez_params({
            "db": db, "id": accession, "rettype": "fasta", "retmode": "text"
        }), timeout=20)
        time.sleep(REQUEST_DELAY)
        if r.status_code != 200 or not r.text.startswith(">"):
            return None
        lines = r.text.strip().splitlines()
        return "".join(lines[1:]).upper()
    except Exception:
        return None


def run_online_checks(row):
    results = {}

    ncbi_acc = clean(row["NCBI_Accession"])
    prot_acc = clean(row["Protein_accession"])
    full = clean(row["Full_Sequence"]) if not pd.isna(row["Full_Sequence"]) else None
    cds = clean(row["CDS_Sequence"])
    prot_stated = clean(row["Protein_Sequence"])
    ipr_stated = clean(row["InterPro_Domain"])

    # --- nucleotide accession ---
    if ncbi_acc:
        seq = fetch_ncbi_sequence(ncbi_acc, "nuccore")
        if seq is None:
            results["NCBI_Accession_lookup"] = ("FAIL", f"{ncbi_acc} not found / not fetchable on NCBI Nucleotide")
        else:
            if full and seq == full:
                results["NCBI_Accession_lookup"] = ("PASS", "matches Full_Sequence exactly")
            elif full and (seq in full or full in seq):
                results["NCBI_Accession_lookup"] = ("WARN", "sequence overlaps but is not an exact match to Full_Sequence")
            elif cds and cds in seq:
                results["NCBI_Accession_lookup"] = ("WARN", "accession resolves and contains the CDS, but doesn't match Full_Sequence (may be CDS-only record)")
            else:
                results["NCBI_Accession_lookup"] = ("FAIL", "accession resolves but sequence doesn't match Full_Sequence or CDS_Sequence")

    # --- protein accession ---
    if prot_acc:
        seq = fetch_ncbi_sequence(prot_acc, "protein")
        if seq is None:
            results["Protein_accession_lookup"] = ("FAIL", f"{prot_acc} not found / not fetchable on NCBI Protein")
        else:
            seq_nostop = seq.rstrip("*")
            if prot_stated and seq_nostop == prot_stated:
                results["Protein_accession_lookup"] = ("PASS", "matches Protein_Sequence exactly")
            else:
                results["Protein_accession_lookup"] = ("FAIL", f"NCBI protein sequence (len {len(seq_nostop)}) != stated Protein_Sequence (len {len(prot_stated) if prot_stated else 0})")

    # --- Gene name / species vs NCBI record title ---
    gene_name = row["Gene_Name"]
    abbrev = row.get("Abbreviation")
    species = str(row["Species"]).strip() if not pd.isna(row["Species"]) else None
    if ncbi_acc:
        summary = fetch_ncbi_summary(ncbi_acc, "nuccore")
        if summary is None:
            results["Gene_Name_vs_NCBI"] = ("SKIP", f"couldn't fetch NCBI summary for {ncbi_acc}")
        else:
            title, organism = summary["title"], summary["organism"]
            if name_matches_title(gene_name, abbrev, title):
                results["Gene_Name_vs_NCBI"] = ("PASS", f"'{gene_name}'/'{abbrev}' found in NCBI title: \"{title}\"")
            else:
                results["Gene_Name_vs_NCBI"] = ("FAIL", f"neither '{gene_name}' nor '{abbrev}' found in NCBI title: \"{title}\"")

            if species_matches_organism(species, organism):
                results["Species_vs_NCBI"] = ("PASS", "")
            else:
                results["Species_vs_NCBI"] = ("FAIL", f"table says '{species}', NCBI record organism is '{organism}'")

    # --- InterPro via UniProt cross-reference (best effort) ---
    if prot_acc and ipr_stated:
        try:
            r = requests.get(f"{UNIPROT_ENTRY}/search", params={
                "query": f"xref:refseq-{prot_acc}", "fields": "accession,xref_interpro", "format": "json"
            }, timeout=20)
            time.sleep(REQUEST_DELAY)
            data = r.json()
            hits = data.get("results", [])
            if not hits:
                results["InterPro_Domain"] = ("SKIP", f"no UniProt entry cross-references RefSeq {prot_acc}")
            else:
                uni_acc = hits[0]["primaryAccession"]
                xrefs = hits[0].get("uniProtKBCrossReferences", [])
                ipr_ids = [x["id"] for x in xrefs if x.get("database") == "InterPro"]
                if ipr_stated in ipr_ids:
                    results["InterPro_Domain"] = ("PASS", f"confirmed via UniProt {uni_acc}")
                elif ipr_ids:
                    results["InterPro_Domain"] = ("FAIL", f"UniProt {uni_acc} lists {ipr_ids}, not {ipr_stated}")
                else:
                    results["InterPro_Domain"] = ("SKIP", f"UniProt {uni_acc} found but has no InterPro cross-references")
        except Exception as e:
            results["InterPro_Domain"] = ("SKIP", f"lookup error: {e}")

    return results


# ------------------------- main -------------------------

def main():
    print(f"Reading {SHEET_NAME!r} from {INPUT_XLSX} ...")
    df = pd.read_excel(INPUT_XLSX, sheet_name=SHEET_NAME)
    print(f"{len(df)} rows loaded.")

    all_rows = []
    for i, row in df.iterrows():
        rid, gene = row["Record_ID"], row["Gene_Name"]
        print(f"[{i+1}/{len(df)}] Record {rid} ({gene}) ...")

        checks = run_offline_checks(row)
        if RUN_ONLINE_CHECKS:
            try:
                checks.update(run_online_checks(row))
            except Exception as e:
                print(f"    ! online check error for {gene}: {e}")

        for field, (status, detail) in checks.items():
            all_rows.append({
                "Record_ID": rid, "Gene_Name": gene, "Field": field,
                "Status": status, "Detail": detail
            })

    report = pd.DataFrame(all_rows)
    issues_only = report[report["Status"].isin(["FAIL", "WARN"])]

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Table1_Original", index=False)
        report.to_excel(writer, sheet_name="All_Checks", index=False)
        issues_only.to_excel(writer, sheet_name="Issues_Only", index=False)

    style_workbook(OUTPUT_XLSX)

    n_fail = (report["Status"] == "FAIL").sum()
    n_warn = (report["Status"] == "WARN").sum()
    n_skip = (report["Status"] == "SKIP").sum()
    n_pass = (report["Status"] == "PASS").sum()
    print()
    print(f"Done. PASS={n_pass} FAIL={n_fail} WARN={n_warn} SKIP={n_skip}")
    print(f"Report written to {OUTPUT_XLSX}")


def style_workbook(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.load_workbook(path)
    font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True)
    fills = {
        "FAIL": PatternFill("solid", fgColor="F8CBAD"),
        "WARN": PatternFill("solid", fgColor="FFE699"),
        "SKIP": PatternFill("solid", fgColor="D9D9D9"),
        "PASS": PatternFill("solid", fgColor="C6E0B4"),
    }
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        status_col = None
        for cell in ws[1]:
            cell.font = header_font
            if cell.value == "Status":
                status_col = cell.column
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font
            if status_col:
                status_val = ws.cell(row=row[0].row, column=status_col).value
                if status_val in fills:
                    ws.cell(row=row[0].row, column=status_col).fill = fills[status_val]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)
    wb.save(path)


if __name__ == "__main__":
    if NCBI_EMAIL == "you@example.com":
        print("!! Edit NCBI_EMAIL at the top of this script before running (NCBI requires it). !!")
        sys.exit(1)
    main()
